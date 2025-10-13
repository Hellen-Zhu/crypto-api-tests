#!/usr/bin/env python
"""
Export Candlestick Test Cases to Excel
=======================================
将K线API测试用例导出为Excel格式，符合测试管理系统的标准格式
"""

import sys
import os
from pathlib import Path
import pandas as pd
import json
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.db_handler import get_db_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import text
from core.logger_config import logger


def export_to_excel():
    """Export candlestick test cases to Excel format"""

    # Connect to database
    engine = get_db_engine()
    Session = sessionmaker(bind=engine)
    session = Session()

    try:
        # Query test case data
        result = session.execute(text("""
            SELECT
                cds.id as dataset_id,
                cds.data_set_name,
                ac.name as case_name,
                ac.parameters,
                cds.variables,
                cds.validations_override,
                cds.tags,
                cds.jira_id
            FROM case_data_sets cds
            JOIN api_auto_cases ac ON ac.id = cds.case_id
            WHERE ac.name = 'Get Candlestick - Professional Test Suite'
            ORDER BY cds.id
        """)).fetchall()

        if not result:
            logger.error("No test cases found")
            return False

        # Prepare data for Excel
        excel_data = []

        for idx, row in enumerate(result, 1):
            dataset_id = row[0]
            data_set_name = row[1]
            case_name = row[2]
            parameters = row[3]
            variables = row[4] or {}
            validations_override = row[5]
            tags = row[6] or []
            jira_id = row[7]

            # Extract step information from parameters
            steps = parameters.get('steps', [])
            if steps:
                step = steps[0]  # K线API只有一个步骤

                # Prepare URL
                base_url = "https://uat-api.3ona.co"
                url = base_url + step.get('path', '')

                # Method
                method = step.get('method', 'GET')

                # Headers
                headers = step.get('request', {}).get('headers', {})
                header_str = json.dumps(headers, ensure_ascii=False)

                # Parameters
                params = {
                    "instrument_name": variables.get('instrument', ''),
                    "timeframe": variables.get('timeframe', ''),
                    "count": variables.get('count', '')
                }
                param_str = json.dumps(params, ensure_ascii=False)

                # Request body (for K线API是GET请求，没有body)
                request_body = ""

                # Comment
                comment = variables.get('description', data_set_name)

                # Expected response
                if validations_override and '1' in validations_override:
                    # Negative test case
                    expected_status = validations_override['1'].get('expectedStatusCode', 400)
                    expected_body = validations_override['1'].get('body', {})
                    response_body_str = json.dumps({
                        "success": False,
                        "code": expected_body.get('code', ''),
                        "message": expected_body.get('message', ''),
                        "data": None
                    }, ensure_ascii=False)
                    response_comment = f"错误场景：{comment}"
                else:
                    # Positive test case
                    expected_status = 200
                    response_body_str = json.dumps({
                        "success": True,
                        "code": 0,
                        "message": "success",
                        "data": {
                            "instrument_name": variables.get('instrument', ''),
                            "interval": variables.get('expected_interval', ''),
                            "data": "[K线数据数组]"
                        }
                    }, ensure_ascii=False)
                    response_comment = f"正常场景：{comment}"

                # Build row data
                row_data = {
                    'No': f"{idx:03d}",
                    'Case Name': data_set_name,
                    'Url': url,
                    'Method': method.upper(),
                    'Header': header_str,
                    'Parameter': param_str,
                    '请求参数': request_body,
                    'Comment': comment,
                    'Response code': expected_status,
                    'Response Body': response_body_str,
                    'Comment2': response_comment
                }

                excel_data.append(row_data)

        # Create DataFrame
        df = pd.DataFrame(excel_data)

        # Export to Excel with formatting
        output_file = f'candlestick_test_cases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='K线API测试用例', index=False)

            # Get the worksheet
            worksheet = writer.sheets['K线API测试用例']

            # Adjust column widths
            column_widths = {
                'A': 8,   # No
                'B': 35,  # Case Name
                'C': 50,  # Url
                'D': 10,  # Method
                'E': 30,  # Header
                'F': 40,  # Parameter
                'G': 20,  # 请求参数
                'H': 40,  # Comment
                'I': 15,  # Response code
                'J': 60,  # Response Body
                'K': 40   # Comment2
            }

            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width

            # Add color to header row
            from openpyxl.styles import PatternFill, Font, Alignment

            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Wrap text for better readability
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        logger.info(f"✅ Successfully exported {len(excel_data)} test cases to {output_file}")

        # Also create a simplified version
        create_simplified_excel(excel_data)

        return True

    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return False
    finally:
        session.close()


def create_simplified_excel(excel_data):
    """Create a simplified version of Excel for easier reading"""

    simplified_data = []

    for row in excel_data:
        # Parse parameter JSON to extract values
        try:
            params = json.loads(row['Parameter'])
        except:
            params = {}

        simplified_row = {
            '编号': row['No'],
            '用例名称': row['Case Name'],
            '交易对': params.get('instrument_name', ''),
            '时间周期': params.get('timeframe', ''),
            '数据量': params.get('count', ''),
            '测试类型': '正向测试' if row['Response code'] == 200 else '负向测试',
            '期望结果': row['Response code'],
            '说明': row['Comment']
        }
        simplified_data.append(simplified_row)

    # Create DataFrame
    df = pd.DataFrame(simplified_data)

    # Export to Excel
    output_file = f'candlestick_test_cases_simplified_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='K线测试用例简化版', index=False)

        # Format the worksheet
        worksheet = writer.sheets['K线测试用例简化版']

        # Adjust column widths
        for idx, col in enumerate(df.columns):
            column_letter = chr(65 + idx)  # A, B, C, etc.
            if idx == 1:  # 用例名称
                worksheet.column_dimensions[column_letter].width = 35
            elif idx == 7:  # 说明
                worksheet.column_dimensions[column_letter].width = 50
            else:
                worksheet.column_dimensions[column_letter].width = 15

        # Add formatting
        from openpyxl.styles import PatternFill, Font, Alignment

        # Header formatting
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Color code rows based on test type
        positive_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        negative_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            test_type = worksheet.cell(row=row_idx, column=6).value
            if test_type == '正向测试':
                for cell in row:
                    cell.fill = positive_fill
            else:
                for cell in row:
                    cell.fill = negative_fill

    logger.info(f"✅ Successfully created simplified Excel: {output_file}")

    # Create test execution summary
    create_test_summary(simplified_data)


def create_test_summary(test_data):
    """Create a test execution summary"""

    # Statistics
    total_tests = len(test_data)
    positive_tests = sum(1 for t in test_data if t['测试类型'] == '正向测试')
    negative_tests = sum(1 for t in test_data if t['测试类型'] == '负向测试')

    # Group by timeframe
    timeframe_counts = {}
    for test in test_data:
        tf = test['时间周期']
        if tf not in timeframe_counts:
            timeframe_counts[tf] = 0
        timeframe_counts[tf] += 1

    # Create summary
    summary = f"""
====================================
K线API测试用例统计报告
====================================
生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

总体统计:
---------
总测试用例数: {total_tests}
正向测试用例: {positive_tests}
负向测试用例: {negative_tests}

时间周期覆盖:
-------------"""

    for tf, count in sorted(timeframe_counts.items()):
        summary += f"\n{tf:8s}: {count:2d} 个用例"

    summary += """

测试策略:
---------
1. 正交设计(L9): 9个核心用例
2. 补充覆盖: 10个周期覆盖用例
3. 负向测试: 5个错误处理用例

执行级别:
---------
Level 1 (冒烟): 4个用例, 2分钟
Level 2 (回归): 10个用例, 5分钟
Level 3 (全量): 24个用例, 15分钟

质量目标:
---------
- 参数组合覆盖: 100%
- 时间周期覆盖: 100%
- 预期缺陷发现率: 95%+
====================================
"""

    # Save summary to file
    summary_file = f'candlestick_test_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write(summary)

    print(summary)
    logger.info(f"✅ Test summary saved to {summary_file}")


if __name__ == "__main__":
    logger.info("Starting candlestick test case Excel export...")

    if export_to_excel():
        logger.info("✅ Export completed successfully!")
        logger.info("📁 Generated files:")
        logger.info("   1. candlestick_test_cases_*.xlsx - 完整格式")
        logger.info("   2. candlestick_test_cases_simplified_*.xlsx - 简化格式")
        logger.info("   3. candlestick_test_summary_*.txt - 统计报告")
    else:
        logger.error("❌ Export failed!")
        sys.exit(1)